#!/usr/bin/env python3
"""
统一数据加载层 - 支持 .csv 和 .xlsx 文件
Priority: 真实国家词表优先架构的数据加载基础设施

支持格式:
  - .csv  → Python 内置 csv 模块
  - .xlsx → pandas.read_excel 或 openpyxl（按优先级尝试）
"""

from __future__ import annotations

import csv
import os
import sys
from typing import Any, Dict, List
from pathlib import Path


# ============================================================
# 依赖检查
# ============================================================

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover - covered by runtime error message
    openpyxl = None  # type: ignore
    _OPENPYXL_IMPORT_ERROR = exc
else:
    _OPENPYXL_IMPORT_ERROR = None


# ============================================================
# 核心加载函数
# ============================================================

def load_table(file_path: str) -> List[Dict[str, Any]]:
    """
    统一加载入口，根据扩展名分派到对应加载器。

    Args:
        file_path: 文件路径

    Returns:
        列表 of 行字典

    Raises:
        FileNotFoundError: 文件不存在
        RuntimeError: 文件格式不支持或依赖缺失
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件不存在: {file_path}")

    ext = Path(file_path).suffix.lower()

    if ext == ".csv":
        return _load_csv(file_path)
    elif ext == ".xlsx":
        return _load_xlsx(file_path)
    else:
        raise RuntimeError(f"不支持的文件格式: {ext}，仅支持 .csv / .xlsx")


def _load_csv(file_path: str) -> List[Dict[str, Any]]:
    """使用 Python 内置 csv 模块加载 CSV"""
    data = []
    with open(file_path, "r", encoding="utf-8") as f:
        # 推断分隔符
        sample = f.read(1024)
        f.seek(0)
        delimiter = "," if "\t" not in sample or sample.count("\t") < sample.count(",") else "\t"
        reader = csv.DictReader(f, delimiter=delimiter)
        for row in reader:
            cleaned = {k.strip(): v.strip() if v else "" for k, v in row.items() if k}
            data.append(cleaned)
    return data


def _load_xlsx(file_path: str) -> List[Dict[str, Any]]:
    """使用 openpyxl 加载 xlsx；缺失依赖时抛出明确提示"""
    if openpyxl is None:
        raise RuntimeError(
            "读取 .xlsx 文件需要 openpyxl。请运行 'pip install openpyxl' 后重试。"
        ) from _OPENPYXL_IMPORT_ERROR

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(headers)]
    data: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(cell is not None for cell in row):
            continue
        record: Dict[str, Any] = {}
        for i, cell in enumerate(row):
            key = headers[i] if i < len(headers) else f"col_{i}"
            if isinstance(cell, str):
                record[key] = cell.strip()
            else:
                record[key] = cell if cell is not None else ""
        data.append(record)
    return data


# ============================================================
# 字段标准化映射
# ============================================================

# 关键词表字段映射
KEYWORD_FIELD_MAP: Dict[str, List[str]] = {
    "keyword":        ["关键词", "keyword", "search_term", "Search Term"],
    "search_volume":  ["月搜索量", "search_volume", "volume", "月搜索", "searches"],
    "conversion_rate": ["购买率", "conversion_rate", "cvr", "转化率", "purchaseRate"],
    "avg_price":      ["均价", "avg_price", "price", "平均价格"],
    "monthly_purchases": ["购买量", "monthly_purchases", "purchases"],
    "click_share":     ["点击份额", "click_share"],
    "avg_cpc":        ["平均点击成本", "avg_cpc", "PPC价格", "PPC竞价", "bid"],
    "spr":            ["SPR", "spr"],
    "title_density":  ["标题密度", "title_density", "titleDensity"],
    "click_concentration": ["点击集中度", "click_concentration"],
    "conv_concentration": ["转化集中度", "conv_concentration"],
    "ac_recommend":   ["AC推荐词", "AC推荐", "ac_recommend"],
    "country":        ["国家", "Country", "country", "market"],
    "model":          ["型号", "model", "Model"],
    "tags":           ["标签", "tags"],
    "product_count":  ["商品数", "product_count", "商品数"],
    "rating_value":   ["评分值", "rating_value", "评分"],
}


def standardize_row(row: Dict[str, Any], field_map: Dict[str, List[str]]) -> Dict[str, Any]:
    """将原始行字段名标准化"""
    standardized = {}
    for std_name, possible_names in field_map.items():
        for name in possible_names:
            if name in row:
                val = row[name]
                standardized[std_name] = val
                break
    # 保留原行中未映射的字段（去重）
    for k, v in row.items():
        if k not in standardized:
            standardized[k] = v
    return standardized


def standardize_keywords(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """标准化关键词表行"""
    result = []
    for row in rows:
        std = standardize_row(row, KEYWORD_FIELD_MAP)
        # 清洗数值字段
        for num_field in ["search_volume", "conversion_rate", "avg_price",
                          "monthly_purchases", "click_share", "avg_cpc",
                          "spr", "title_density", "click_concentration",
                          "conv_concentration", "product_count", "rating_value"]:
            if num_field in std and std[num_field]:
                cleaned = str(std[num_field]).replace(",", "").replace("%", "").replace(" ", "")
                try:
                    if "." in cleaned:
                        std[num_field] = float(cleaned)
                    else:
                        std[num_field] = int(float(cleaned))
                except (ValueError, TypeError):
                    std[num_field] = 0
        result.append(std)
    return result


# ============================================================
# 行数检查（用于快速验证）
# ============================================================

def count_rows(file_path: str) -> int:
    """快速统计文件行数（不含表头）"""
    try:
        rows = load_table(file_path)
        return len(rows)
    except Exception:
        return 0


# ============================================================
# CLI 测试入口
# ============================================================

if __name__ == "__main__":
    import json

    caps = _get_loader_caps()
    print("=" * 50)
    print("统一数据加载层 - 环境检查")
    print("=" * 50)
    print(f"  CSV 支持:       {caps['csv']}")
    print(f"  XLSX(openpyxl): {caps['xlsx_openpyxl']}")
    print(f"  XLSX(pandas):   {caps['xlsx_pandas']}")
    print()

    if len(sys.argv) < 2:
        print("用法: python data_loader.py <file_path>")
        sys.exit(0)

    file_path = sys.argv[1]
    print(f"加载文件: {file_path}")
    try:
        rows = load_table(file_path)
        print(f"  行数: {len(rows)}")
        if rows:
            print(f"  列名: {list(rows[0].keys())}")
            print(f"  示例行: {json.dumps(rows[0], ensure_ascii=False)[:200]}")
    except Exception as e:
        print(f"  错误: {e}")
