#!/usr/bin/env python3
"""SellerSprite / PPC keyword table parser for human-in-the-loop feedback."""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any, Dict, List, Tuple

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - optional dependency in local env
    load_workbook = None

from modules.keyword_utils import GLOBAL_BRAND_BLOCKLIST, is_blocklisted_brand

COLUMN_ALIASES = {
    "keyword": ["keyword", "关键词", "search term", "search_query", "term"],
    "source": ["source", "traffic source", "流量来源", "来源", "channel", "ad type"],
    "search_volume": ["search volume", "月搜索量", "volume", "searches", "搜索量"],
    "conversion": ["conversion", "conversion rate", "cv", "转化率", "购买率", "order rate"],
    "orders": ["orders", "订单量", "sales", "销量", "purchases"],
}


def _normalize_header(value: str) -> str:
    return " ".join(str(value or "").strip().lower().replace("_", " ").split())


def _resolve_columns(columns: List[str]) -> Dict[str, str]:
    normalized = {_normalize_header(col): col for col in columns}
    resolved: Dict[str, str] = {}
    for target, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            col = normalized.get(_normalize_header(alias))
            if col:
                resolved[target] = col
                break
    return resolved


def _to_float(value: Any) -> float:
    if value is None:
        return 0.0
    text = str(value).strip().replace(",", "")
    if not text:
        return 0.0
    try:
        return float(text.replace("%", ""))
    except ValueError:
        return 0.0


def _infer_source(raw_source: str, keyword: str) -> str:
    text = f"{raw_source} {keyword}".lower()
    if any(token in text for token in ["organic", "自然", "seo"]):
        return "organic"
    if any(token in text for token in ["ppc", "sp", "sponsored", "广告"]):
        return "sp"
    return "unknown"


def _has_garbled_text(keyword: str) -> bool:
    text = (keyword or "").strip()
    if not text:
        return True
    weird_chars = sum(1 for ch in text if ord(ch) < 32)
    return weird_chars > 0 or "???" in text or len(text) <= 1


def _suggest_slot(source: str, keyword: str) -> str:
    lowered = (keyword or "").lower()
    token_count = len(lowered.split())
    if source == "organic":
        return "title" if token_count <= 3 else "bullets"
    if source == "sp":
        return "bullets" if token_count <= 4 else "backend"
    return "backend"


def _risk_flag(keyword: str) -> Tuple[str, str, bool]:
    lowered = (keyword or "").lower()
    matched_brand = is_blocklisted_brand(lowered)
    if matched_brand:
        return "blocked_brand", f"命中竞品词 {matched_brand}", False
    if _has_garbled_text(keyword):
        return "garbled", "疑似乱码或低质量词", False
    if any(term in lowered for term in ["free", "best", "cheap"]):
        return "review", "需要人工确认是否适合进入前台文案", True
    return "ok", "", True


def parse_keyword_feedback_table(file_path: str) -> Dict[str, Any]:
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(file_path)

    rows_raw: List[Dict[str, Any]] = []
    if path.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
        if load_workbook is None:
            raise RuntimeError("openpyxl not available for xlsx parsing")
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        iterator = ws.iter_rows(values_only=True)
        headers = [str(cell or "").strip() for cell in next(iterator, [])]
        for row in iterator:
            rows_raw.append({headers[idx]: row[idx] for idx in range(min(len(headers), len(row)))})
    else:
        with open(path, "r", encoding="utf-8", newline="") as fh:
            reader = csv.DictReader(fh)
            headers = list(reader.fieldnames or [])
            rows_raw.extend(list(reader))

    resolved = _resolve_columns(headers)
    keyword_col = resolved.get("keyword")
    if not keyword_col:
        raise ValueError("未识别关键词列，当前不符合 SellerSprite/PPC 词表格式")

    rows: List[Dict[str, Any]] = []
    for row in rows_raw:
        keyword = str(row.get(keyword_col, "") or "").strip()
        if not keyword:
            continue
        raw_source = str(row.get(resolved.get("source", ""), "") or "").strip() if resolved.get("source") else ""
        source = _infer_source(raw_source, keyword)
        search_volume = _to_float(row.get(resolved.get("search_volume", ""), 0)) if resolved.get("search_volume") else 0.0
        conversion = _to_float(row.get(resolved.get("conversion", ""), 0)) if resolved.get("conversion") else 0.0
        orders = _to_float(row.get(resolved.get("orders", ""), 0)) if resolved.get("orders") else 0.0
        risk_flag, reason, keep = _risk_flag(keyword)
        rows.append({
            "keep": keep,
            "keyword": keyword,
            "source": source,
            "raw_source": raw_source,
            "search_volume": search_volume,
            "conversion": conversion,
            "orders": orders,
            "suggested_slot": _suggest_slot(source, keyword),
            "risk_flag": risk_flag,
            "reason": reason,
        })

    rows.sort(key=lambda item: (0 if item["source"] == "organic" else 1, -item["search_volume"], -item["conversion"]))
    return {
        "file_path": str(path),
        "columns": resolved,
        "rows": rows,
        "summary": {
            "total": len(rows),
            "organic": sum(1 for row in rows if row["source"] == "organic"),
            "sp": sum(1 for row in rows if row["source"] == "sp"),
            "blocked": sum(1 for row in rows if row["risk_flag"] != "ok"),
        },
    }


__all__ = ["parse_keyword_feedback_table"]
