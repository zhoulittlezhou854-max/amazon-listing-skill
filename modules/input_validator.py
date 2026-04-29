#!/usr/bin/env python3
"""Non-blocking validation for listing input tables."""

from __future__ import annotations

import csv
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional

try:
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover
    load_workbook = None


REQUIRED_COLUMNS: Dict[str, List[str]] = {
    "attribute_table": ["Field_Name", "Value"],
    "keyword_table": ["keyword", "search_volume"],
    "review_table": ["ASIN", "Bullet_1", "BSR_Rank"],
    "aba_merged": ["keyword", "search_volume"],
}
ALIASED_COLUMNS: Dict[str, Dict[str, List[str]]] = {
    "keyword_table": {
        "keyword": ["keyword", "关键词"],
        "search_volume": ["search_volume", "月搜索量"],
    },
    "aba_merged": {
        "keyword": ["keyword", "关键词"],
        "search_volume": ["search_volume", "月搜索量"],
    },
}
REVIEW_TABLE_SCHEMAS: List[List[str]] = [
    ["ASIN", "Bullet_1", "BSR_Rank"],
    ["ASIN", "Bullet_1", "Bullet_2"],
    ["ASIN", "Data_Type", "Field_Name", "Content_Text"],
]
REQUIRED_NUMERIC_COLUMNS: Dict[str, List[str]] = {
    "keyword_table": ["search_volume"],
    "review_table": ["BSR_Rank"],
    "aba_merged": ["search_volume"],
}


@dataclass
class ValidationWarning:
    table: str
    severity: str
    message: str


def _coerce_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def _resolve_table_path(run_config: Any, table_name: str) -> Optional[Path]:
    input_files = getattr(run_config, "input_files", None) or {}
    path = input_files.get(table_name)
    if not path:
        return None
    return Path(path)


def _read_header(path: Path) -> List[str]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"} and load_workbook is not None:
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.active
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
            return [str(cell).strip() for cell in first_row if cell is not None and str(cell).strip()]
        finally:
            wb.close()
    if suffix == ".txt":
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        first_row = next(reader, [])
    return [str(cell).strip() for cell in first_row if str(cell).strip()]


def _read_sample_rows(path: Path, limit: int = 5) -> List[Dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"} and load_workbook is not None:
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.active
            rows = ws.iter_rows(values_only=True)
            header_row = next(rows, ())
            headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
            samples: List[Dict[str, Any]] = []
            for row in rows:
                payload = {}
                for idx, header in enumerate(headers):
                    if not header:
                        continue
                    payload[header] = row[idx] if idx < len(row) else None
                if any(value not in {None, ""} for value in payload.values()):
                    samples.append(payload)
                if len(samples) >= limit:
                    break
            return samples
        finally:
            wb.close()
    if suffix == ".txt":
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        samples = []
        for row in reader:
            samples.append(dict(row))
            if len(samples) >= limit:
                break
    return samples


def _column_aliases(table_name: str, column: str) -> List[str]:
    return ALIASED_COLUMNS.get(table_name, {}).get(column, [column])


def _resolve_header_name(header: List[str], table_name: str, column: str) -> Optional[str]:
    for alias in _column_aliases(table_name, column):
        if alias in header:
            return alias
    return None


def _missing_required_columns(table_name: str, header: List[str], required_cols: List[str]) -> List[str]:
    if table_name == "review_table":
        for schema in REVIEW_TABLE_SCHEMAS:
            if all(col in header for col in schema):
                return []
        return required_cols
    return [col for col in required_cols if _resolve_header_name(header, table_name, col) is None]


def validate_input_tables(run_config: Any) -> List[ValidationWarning]:
    warnings: List[ValidationWarning] = []
    for table_name, required_cols in REQUIRED_COLUMNS.items():
        path = _resolve_table_path(run_config, table_name)
        if not path or not path.exists():
            warnings.append(
                ValidationWarning(
                    table=table_name,
                    severity="medium",
                    message=f"{table_name} 文件不存在，将使用降级策略",
                )
            )
            continue
        try:
            header = _read_header(path)
        except Exception as exc:
            warnings.append(
                ValidationWarning(
                    table=table_name,
                    severity="high",
                    message=f"{table_name} 读取失败：{exc}",
                )
            )
            continue
        if not header:
            # Legacy text tables are still accepted by the current pipeline.
            continue
        missing = _missing_required_columns(table_name, header, required_cols)
        if missing:
            warnings.append(
                ValidationWarning(
                    table=table_name,
                    severity="high",
                    message=f"{table_name} 缺少必填列：{missing}",
                )
            )
            continue
        numeric_columns = REQUIRED_NUMERIC_COLUMNS.get(table_name) or []
        if not numeric_columns:
            continue
        try:
            sample_rows = _read_sample_rows(path)
        except Exception as exc:
            warnings.append(
                ValidationWarning(
                    table=table_name,
                    severity="high",
                    message=f"{table_name} 示例数据读取失败：{exc}",
                )
            )
            continue
        for column in numeric_columns:
            header_column = _resolve_header_name(header, table_name, column)
            if not header_column:
                continue
            bad_examples: List[str] = []
            for row in sample_rows:
                value = row.get(header_column)
                if value in {None, ""}:
                    bad_examples.append("<empty>")
                elif _coerce_float(value) is None:
                    bad_examples.append(str(value))
                if len(bad_examples) >= 3:
                    break
            if bad_examples:
                warnings.append(
                    ValidationWarning(
                        table=table_name,
                        severity="high",
                        message=(
                            f"{table_name} 列 `{column}` 应为数值类型，"
                            f"当前检测到异常值：{bad_examples}"
                        ),
                    )
                )
    return warnings


def warnings_as_dicts(warnings: List[ValidationWarning]) -> List[Dict[str, Any]]:
    return [asdict(item) for item in warnings]
